from openpyxl import load_workbook, Workbook
import math

def parse_date(date_str):

    date_time = date_str.split()
    if date_time[0].find(':') != -1:
        temp = date_time[0]
        date_time[0] = date_time[1]
        date_time[1] = temp

    date, time = date_time

    if date.find('/') != -1:
        month, day, year = date.split('/')
    elif date.find('-') != -1:
        year, month, day = date.split('-')

    if len(day) == 1:
        day = '0' + day

    if len(month) == 1:
        month = '0' + month

    if len(year) == 2:
        year = '20' + year

    return day, month, year

def distance(location1, location2):
    return int(80 * math.sqrt((location1[0] - location2[0]) ** 2 + (location1[1] - location2[1]) ** 2))


class Time:

    def __init__(self, hour, minutes):

        self.hour = hour
        self.minutes = minutes

    def add(self, other):
        other_hour = other.get_hour()
        other_minutes = other.get_minutes()
        combined_minutes = self.minutes + other_minutes
        added_hours = combined_minutes // 60 + other_hour
        added_minutes = combined_minutes % 60

        return Time(self.hour + added_hours, self.minutes + added_minutes)

    def subtract(self, other):
        if not self.greater(other) and not self.equals(other):
            return -1

        other_hour = other.get_hour()
        other_minutes = other.get_minutes()
        difference_minutes = self.minutes - other_minutes
        resultant_minutes = difference_minutes
        resultant_hours = self.hour - other_hour

        if difference_minutes < 0:
            resultant_minutes += 60
            resultant_hours -= 1

        return Time(resultant_hours, resultant_minutes)

    def copy(self):
        return Time(self.hour, self.minutes)

    def greater(self, other):
        return self.get_time() > other.get_time()

    def equals(self, other):
        return self.get_time() == other.get_time()

    def get_time(self):
        return self.hour, self.minutes

    def get_hour(self):
        return self.hour

    def get_minutes(self):
        return self.minutes

ZERO_TIME = Time(0, 0)
MORNING_STARTIME = Time(4, 0)
MORNING_ENDTIME = Time(14, 0)
EVENING_STARTIME = Time(14, 0)
EVENING_ENDTIME = Time(24, 0)

SHIFTS = {"Morning" : (MORNING_STARTIME, MORNING_ENDTIME), "Evening" : (EVENING_STARTIME, EVENING_ENDTIME)}

class Schedule:

    def __init__(self, begin_shift, end_shift, default_location):


        self.begin_shift = begin_shift
        self.end_shift = end_shift
        self.shift_time = self.end_shift.subtract(self.begin_shift)
        self.schedule = list()
        self.schedule.append(list())
        self.schedule_days = 0
        self.default_location = default_location
        self.recent_location = default_location

    def add_task(self, order, time, order_location):

        if not self.is_available():
            return False

        if self.schedule_days == 1:
            last_end_time = self.schedule[0][-1][2] + movement_time

        else:
            last_end_time = self.begin_shift

        travel_distance = distance(self.recent_location, order_location)

        travel_time = Time(0, 2 * travel_distance)

        start_time = last_end_time.add(travel_time)

        remaining_time = self.end_shift.subtract(start_time)

        if not time.greater(remaining_time):
            self.schedule[0].append((order, start_time, start_time.add(time)))

        else:
            overflow_time = time.subtract(remaining_time)
            self.schedule[0].append((order, start_time, self.end_shift))
            day = 1
            while overflow_time.greater(ZERO_TIME):
                if day >= len(self.schedule):
                    self.schedule.append(list())
                if overflow_time.greater(self.shift_time):
                    self.schedule[day].append((order, self.begin_shift, self.end_shift))
                    overflow_time = overflow_time.subtract(self.shift_time)
                else:
                    self.schedule[day].append((order, self.begin_shift, self.begin_shift.add(overflow_time)))
                    overflow_time = ZERO_TIME
                day += 1

        self.recent_location = order_location

        return True


    def is_available(self):
        return self.schedule_days < 2

    def get_next_start_time(self, current_location, order_location):

        if not self.schedule[0]:
            return self.begin_shift

        travel_distance = distance(current_location, order_location)
        travel_time = Time(0, 2 * travel_distance)

        return self.schedule[0][-1][2].add(travel_time)


    def get_time_remaining(self):
        if not self.schedule[0]:
            return self.shift_time
        else:
            return self.end_shift.subtract(self.schedule[0][-1][2])

    def update(self):
        self.schedule.pop(0)
        if not self.schedule:
            self.schedule.append(list())
        self.recent_location = self.default_location

    def get_recent_location(self):
        return self.recent_location

    def print_today(self):
        for order_information in self.schedule[0]:
            order, start_time, end_time = order_information
            print(order.get_key(), start_time.get_time(), end_time.get_time())


class Facility_Schedule:

    def __init__(self, max_capacity):

        self.daytime = EVENING_ENDTIME.subtract(MORNING_STARTIME)
        self.schedule = list()
        self.schedule.append(list())
        self.schedule[0].append((0, EVENING_ENDTIME))
        self.max_capacity = max_capacity

    def get_start_times(self, duration):

        valid_intervals = []
        is_valid = False
        start_time = None
        prev_endtime = MORNING_STARTIME

        for capacity, end_time in self.schedule[0]:
            if is_valid and capacity >= self.max_capacity:
                valid_intervals.append((start_time, prev_endtime))
                is_valid = False
            if not is_valid and capacity < self.max_capacity:
                start_time = prev_endtime
                is_valid = True

            prev_endtime = end_time

        if is_valid:
            valid_intervals.append((start_time, prev_endtime))

        valid_start_times = []
        for start_time, end_time in valid_intervals:

            if end_time.subtract(start_time).greater(duration):
                valid_start_times.append((start_time, end_time.subtract(duration)))

        return valid_start_times


    def add_task(self, order, start_time, time):

        end_time = start_time.add(time)

        index = 0
        start_index = -1
        while index < len(self.schedule[0]):
            capacity, end_interval = self.schedule[0][index]
            if end_interval.greater(start_time):
                start_index = index
                break
            index += 1

        index = start_index
        while index < len(self.schedule[0]):
            capacity, end_interval = self.schedule[0][index]
            if capcity == self.max_capacity:
                return False
            if end_interval.greater(end_time):
                break

            index += 1

        end_index = index

        new_schedule = list()

        if start_index != -1:
            for index in range(start_index + 1):
                capacity, end_interval = self.schedule[0][index]
                new_schedule.append((capacity, end_interval.copy()))

        if not start_time.equals(self.schedule[0][start_index]):
            new_schedule.append((capacity + 1, start_time))

        for index in range(start_index + 1, end_index):
            capacity, end_interval = self.schedule[0][index]
            new_schedule.append((capacity + 1, end_interval.copy()))

        if not end_time.greater(EVENING_ENDTIME):
            if not end_time.equals(self.schedule[0][end_index - 1]):
                new_schedule.append((capacity, end_time))

            for index in range(end_index, len(self.schedule[0])):
                capacity, end_interval = self.schedule[0][index]
                new_schedule.append((capacity, end_interval.copy()))

        else:

            self.schedule[0] = new_schedule
            overflow_time = end_time.subtract(EVENING_ENDTIME)

            day = 1
            while overflow_time.greater(ZERO_TIME):
                if day >= len(self.schedule):
                    self.schedule.append(list())
                if overflow_time.greater(self.daytime):
                    self.schedule[day].append((1, EVENING_ENDTIME))
                    overflow_time = overflow_time.subtract(daytime)
                else:
                    self.schedule[day].append((1, MORNING_STARTIME.add(overflow_time)))
                    overflow_time = ZERO_TIME
                day += 1

        return True


class Read_Data():

    def __init__(self):

        self.NULL = {None, "=""", ""}
        self.WORKFILE = "data_file.xlsx"
        self.SHEET_NAMES = ["Equipment Details", "Worker Details", "Facility Details", "Work Order Examples"]
        self.wb = load_workbook(self.WORKFILE)

        self.EQUIPMENT_FIELDS = ["failure", "fix", "fac1", "fac2", "fac3", "fac4", "fac5"]
        self.WORKER_FIELDS = ["certifications", "shifts"]
        self.FACILITY_FIELDS = ["latitude", "longitude", "occupancy"]
        self.ORDER_FIELDS = ["facility", "type", "id", "priority", "completion_time", "submission_time"]

        self.EQUIPMENT = self.read_sheet(self.SHEET_NAMES[0], 3, 2, self.EQUIPMENT_FIELDS)
        self.WORKER = self.read_sheet(self.SHEET_NAMES[1], 2, 2, self.WORKER_FIELDS)
        self.FACILITY = self.read_sheet(self.SHEET_NAMES[2], 3, 2, self.FACILITY_FIELDS)
        self.ORDER = self.read_sheet(self.SHEET_NAMES[3], 3, 2, self.ORDER_FIELDS)

        self.DISTANCE_MATRIX = self.generate_distance_matrix()


    def read_sheet(self, sheetname, start_row, start_col, fields):

        data = dict()
        ws = self.wb[sheetname]
        row = start_row

        while True:
            key = ws.cell(row, start_col).value
            if key in self.NULL:
                break
            field_dict = dict()
            col = start_col + 1
            for index, field in enumerate(fields):
                field_dict[field] = str(ws.cell(row, col + index).value)
            data[key] = field_dict

            row += 1

        return data

    def generate_distance_matrix(self):

        n = len(self.FACILITY)
        coordinates = list()
        for fields in self.FACILITY.values():
            pair = (float(fields["latitude"]), float(fields["longitude"]))
            coordinates.append(pair)

        matrix = list()
        for index1 in range(n):
            distance = list()
            latitude1 = coordinates[index1][0]
            longitude1 = coordinates[index1][1]
            for index2 in range(n):
                latitude2 = coordinates[index2][0]
                longitude2 = coordinates[index2][1]
                distance.append(math.sqrt((latitude2 - latitude1) ** 2 + (longitude2 - longitude1) ** 2))

            matrix.append(distance)

        return matrix


    def get_equipment(self):
        return self.EQUIPMENT

    def get_worker(self):
        return self.WORKER

    def get_facility(self):
        return self.FACILITY

    def get_order(self):
        return self.ORDER


class Equipment():

    def __init__(self, key, equipment_fields):

        self.key = key
        self.failure = equipment_fields["failure"]
        self.fix = equipment_fields["fix"]
        self.fac1 = equipment_fields["fac1"]
        self.fac2 = equipment_fields["fac2"]
        self.fac3 = equipment_fields["fac3"]
        self.fac4 = equipment_fields["fac4"]
        self.fac5 = equipment_fields["fac5"]

    def get_key(self):
        return self.key

    def get_failure(self):
        return self.failure

    def get_fix(self):
        return self.fix

    def get_fac1(self):
        return self.fac1

    def get_fac2(self):
        return self.fac2

    def get_fac3(self):
        return self.fac3

    def get_fac4(self):
        return self.fac4

    def get_fac5(self):
        return self.fac5


class Worker():

    def __init__(self, key, worker_fields, default_location):

        self.key = key
        certifications = worker_fields["certifications"]
        self.certifications = set(certifications.split(", "))
        self.shifts = worker_fields["shifts"]

        self.start_shift, self.end_shift = SHIFTS[self.shifts]
        self.schedule = Schedule(self.start_shift, self.end_shift, default_location)

    def get_key(self):
        return self.key

    def get_certifications(self):
        return self.certifications

    def get_shifts(self):
        return self.shifts

    def get_location(self):
        return self.schedule.get_recent_location()

    def get_schedule(self):
        return self.schedule


class Facility():

    def __init__(self, key, facility_fields):

        self.key = key
        self.latitude = float(facility_fields["latitude"])
        self.longitude = float(facility_fields["longitude"])
        self.occupancy = int(facility_fields["occupancy"])

        self.schedule = Facility_Schedule(self.occupancy)

    def get_key(self):
        return self.key

    def get_location(self):
        return self.latitude, self.longitude

    def get_latitude(self):
        return self.latitude

    def get_longitude(self):
        return self.longitude

    def get_occupancy(self):
        return self.occupancy

    def get_schedule(self):
        return self.schedule


class Order():

    def __init__(self, key, order_fields):

        self.key = key
        self.facility = order_fields["facility"]
        self.type = order_fields["type"]
        self.id = order_fields["id"]
        self.priority = int(order_fields["priority"])
        self.completion_time = int(order_fields["completion_time"])
        self.submission_time = order_fields["submission_time"]

    def get_key(self):
        return self.key

    def get_facility(self):
        return self.facility

    def get_type(self):
        return self.type

    def get_id(self):
        return self.id

    def get_priority(self):
        return self.priority

    def get_completion_time(self):
        return self.completion_time

    def get_submission_time(self):
        return self.submission_time



class Data():

    def __init__(self):

        self.data = Read_Data()
        self.equipment_data = self.data.get_equipment()
        self.worker_data = self.data.get_worker()
        self.facility_data = self.data.get_facility()
        self.order_data = self.data.get_order()

        self.equipment_objs = dict()
        self.worker_objs = dict()
        self.facility_objs = dict()
        self.order_objs = dict()

        for key, equipment_fields in self.equipment_data.items():
            self.equipment_objs[key] = Equipment(key, equipment_fields)

        for key, facility_fields in self.facility_data.items():
            self.facility_objs[key] = Facility(key, facility_fields)

        count = 0
        sum_latitiudes = 0
        sum_longitudes = 0
        for facility in self.facility_objs.values():
            sum_latitiudes += facility.get_latitude()
            sum_longitudes += facility.get_longitude()
            count += 1
        self.default_location = (sum_latitiudes / count, sum_longitudes / count)

        for key, worker_fields in self.worker_data.items():
            self.worker_objs[key] = Worker(key, worker_fields, self.default_location)

        for key, order_fields in self.order_data.items():
            self.order_objs[key] = Order(key, order_fields)

        distinct_days = set()
        order_days = list()

        for key, order in self.order_objs.items():

            submission_time = order.get_submission_time()
            order_day = int(parse_date(submission_time)[0])
            distinct_days.add(order_day)
            order_days.append((key, order_day))

        min_day = min(distinct_days)
        max_day = max(distinct_days)
        total_days = max_day - min_day + 1

        self.daily_order_objs = list()
        for _ in range(total_days):
            self.daily_order_objs.append(dict())
        for order in order_days:
            key, order_day = order
            day_offset = order_day - min_day
            self.daily_order_objs[day_offset][key] = self.get_order(key)


    def get_equipment(self, name):
        if name in self.equipment_objs:
            return self.equipment_objs[name]
        return None

    def get_worker(self, name):
        if name in self.worker_objs:
            return self.worker_objs[name]
        return None

    def get_facility(self, name):
        if name in self.facility_objs:
            return self.facility_objs[name]
        return None

    def get_order(self, name):
        if name in self.order_objs:
            return self.order_objs[name]
        return None

    def get_all_equipment(self):
        return self.equipment_objs

    def get_all_workers(self):
        return self.worker_objs

    def get_all_facilities(self):
        return self.facility_objs

    def get_all_orders(self):
        return self.order_objs

    def get_all_daily_orders(self):
        return self.daily_order_objs


def main():
    data = Data()
    print(data.get_worker("Bob").get_certifications())

if __name__ == '__main__':
    main()
